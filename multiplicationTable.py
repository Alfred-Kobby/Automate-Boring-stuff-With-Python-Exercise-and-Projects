import sys
import openpyxl


if len(sys.argv) == 2:
    # Get the max size of the N by N matrix
    max_row = int(sys.argv[1])

    # create a workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Loop through to populate sheet
    for i in range(2, max_row+2): #Skip first row hence range starts with 2
        sheet.cell(row=i, column=1).value = i - 1
        row_val = i-1
        for k in range(2, max_row+2): #Skip first column hence range starts with 2
            sheet.cell(row=1, column=i).value = i - 1
            column_val = k-1
            sheet.cell(row=i, column=k).value = row_val * column_val

    # Save the workbook
    wb.save('multiplication table.xlsx')