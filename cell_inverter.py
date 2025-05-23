import openpyxl

wb = openpyxl.load_workbook("items.xlsx")
sheet = wb.active

sb = openpyxl.Workbook()
sheet1 = sb.active

for i in range(1, sheet.max_column+1):
    for k in range(1, sheet.max_row+1):
        sheet1.cell(row=i, column=k).value = sheet.cell(row=k, column=i).value
    print()
sb.save("items1.xlsx")