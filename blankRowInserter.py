import sys
import openpyxl


if len(sys.argv) == 4:

    wb = openpyxl.load_workbook(sys.argv[3])
    sheet = wb.active

    sb = openpyxl.Workbook()
    sheet1 = sb.active

    row = int(sys.argv[1])
    column = int(sys.argv[2])

    #Copy the first rows before inserting the space
    for i in range(1, row):
        for k in range(1, sheet.max_column):
            sheet1.cell(row=i, column=k).value = sheet.cell(row=i, column=k).value

    #insert the remaining rows after the space
    for i in range(row+column, sheet.max_row):
        for k in range(1, sheet.max_column):
            sheet1.cell(row=i, column=k).value = sheet.cell(row=i, column=k).value

    sb.save("myProduce1.xlsx")
else:
    print("Pass right arguments")