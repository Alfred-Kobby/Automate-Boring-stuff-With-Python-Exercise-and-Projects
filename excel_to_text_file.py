import sys
import openpyxl

if len(sys.argv) > 1:
    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active

    for i in range(1, sheet.max_column+1):
        with open(f"newFile{i}.txt", 'a') as file:
            for k in range(1, sheet.max_row+1):
                print(sheet.cell(row=k, column=i).value)
                file.write(sheet.cell(row=k, column=i).value + "\n")
            print("------------------")
        file.close()
else:
    print("No excel file found")
